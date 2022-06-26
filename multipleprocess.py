from concurrent.futures import ProcessPoolExecutor
from openpyxl import load_workbook
import time
from itertools import islice
import pandas as pd
import os

test_file = "C:\\Users\\Lenovo\Desktop\\make money\\拆分\\test\\汇总2.xlsx"


def worksheetToDataFrame(wb, sheetname):

    ws = wb[sheetname]
    data = ws.values
    cols = next(data)[0:]
    data = list(data)
    data = (islice(r, 0, None) for r in data)
    df = pd.DataFrame(data, columns=cols)            
    # remove unnamedcolumn
    df = removeNullColumn(df)

    return df

def removeNullColumn(df):
    cols = [c for c in df.columns.values if c!= None and c != 'Unnamed']
    df = df[cols]
    return df

def parallel_worksheet(sheetname):
    begin = time.time()
    wb = load_workbook(test_file, read_only=True,
                       data_only=True, keep_links=False)
    df = worksheetToDataFrame(wb, sheetname)
    end = time.time()
    print("    {0} {1:.2f}s".format(sheetname, end - begin))
    #wb.close()
    return end-begin


def parallel_read():
    print("Parallised Read")
    begin = time.time()
    wb = load_workbook(test_file, read_only=True,
                       keep_links=False, data_only=True)
    print("    Workbook loaded {0:.2f}s".format(time.time() - begin))
    sheets = ["总表"]
    with ProcessPoolExecutor(4) as pool:
        for spent in pool.map(parallel_worksheet, sheets, chunksize=1):
            print("%2f" %spent)
    end = time.time()
    wb.close()
    print("    Total time {0:.2f}s".format(end - begin))

def serialize_read():
    
    begin = time.time()
    wb = load_workbook(test_file, read_only=True,
                       data_only=True, keep_links=False)
    print("Serialize Read", os.access(test_file, os.R_OK))
    df = worksheetToDataFrame(wb, "个税汇总")
    wb.close()
    end = time.time()
    print("    Total time {0:.2f}s".format(end - begin))

if __name__ == "__main__":
    print(os.cpu_count())
    #parallel_read()
    serialize_read()
    #testJob()
