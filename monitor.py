# -*- coding: utf-8 -*-
"""
Created on Thu Jun 16 17:21:38 2022

Notes:
    Need to install pyinotify first (conda install pyinotify)

@author: Jane Huang
"""

import os
import time
import numpy as np
import pandas as pd
import openpyxl
import logging
import win32api,win32con
from popbubble import show_msg
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles import Border
from openpyxl.styles import Side
from itertools import islice

# 需要手动更新的目录
userPath="C:\\Users\\Lenovo\Desktop\\make money\\拆分\\test"
# 拆分的配置信息
configPath = 'C:\\Users\\Lenovo\Desktop\\make money\\拆分\\config\\configuration.xlsx'
# userPath下所有打开文件
openedFileList = []
# userPath下所有打开文件的初始副本
# key:('C:\Users\汇总.xlsx', sheetname), value:dataframe
dfDict = {} 
# 公司列，索引列，等配置信息
idColumn = 0 # idColumn不用固定，但必须命名为编号
idStr = "编号"
configDF = None
saveCount = 0
# 气泡提示停留时间
leftSeconds = 3

#设置excel样式
cellFont = Font(
    name="宋体",   # 字体
    size=9,         # 字体大小
    color="000000",  # 字体颜色，用16进制rgb表示
    bold=False,       # 是否加粗，True/False
    italic=False,     # 是否斜体，True/False
    #strike=False,     # 是否使用删除线，True/False
    #underline=None,  # 下划线, 可选'singleAccounting', 'double', 'single', 'doubleAccounting'
)
cellAlignment = Alignment(
    horizontal='center',  # 水平对齐，可选general、left、center、right、fill、justify、centerContinuous、distributed
    vertical='center',  # 垂直对齐， 可选top、center、bottom、justify、distributed
    #text_rotation=0,  # 字体旋转，0~180整数
    #wrap_text=False,  # 是否自动换行
    #shrink_to_fit=False,  # 是否缩小字体填充
    #indent=0,  # 缩进值
)
cellBorder = Border(bottom=Side(style='thin'),
                    right=Side(style='thin'))

# log
logger = None

def setLogger():

    # 第一步，创建一个logger
    logger = logging.getLogger()
    if not logger.handlers:
        logger.setLevel(logging.INFO)  # Log等级总开关  此时是INFO
        
        # 第二步，创建一个handler，用于写入日志文件
        logfile = './log.txt'
        fh = logging.FileHandler(logfile, mode='a')  # open的打开模式这里可以进行参考
        fh.setLevel(logging.DEBUG)  # 输出到file的log等级的开关
        
        # 第三步，再创建一个handler，用于输出到控制台
        ch = logging.StreamHandler()
        ch.setLevel(logging.DEBUG)   # 输出到console的log等级的开关
        
        # 第四步，定义handler的输出格式（时间，文件，行数，错误级别，错误提示）
        formatter = logging.Formatter("%(asctime)s - %(filename)s[line:%(lineno)d] - %(levelname)s: %(message)s")
        fh.setFormatter(formatter)
        ch.setFormatter(formatter)
        
        # 第五步，将logger添加到handler里面
        logger.addHandler(fh)
        logger.addHandler(ch)
    
    # 日志级别
    # DEBUG：详细的信息,通常只出现在诊断问题上
    # INFO：确认一切按预期运行
    # WARNING（默认）：一个迹象表明,一些意想不到的事情发生了,或表明一些问题在不久的将来(例如。磁盘空间低”)。这个软件还能按预期工作。
    # ERROR：更严重的问题,软件没能执行一些功能
    # CRITICAL：一个严重的错误,这表明程序本身可能无法继续运行
    return logger


def changeColumnToNum(column):
    index = 0
    column = column.strip().upper()
    if len(column) > 1:
        base = 1
        for char in column[::-1]:
            index += base * (ord(char) - ord('A') + 1)
            base = base * 23
        index -= 1
    else:
        index = ord(column) - ord('A')
    return index

def removeNullColumn(df):
    cols = [c for c in df.columns.values if c!= None and c != 'Unnamed']
    df = df[cols]
    return df

def worksheetToDataFrame(wb, sheetname, ws=None):

    if ws == None:
        ws = wb[sheetname]
    data = ws.values
    cols = next(data)[0:]
    data = list(data)
    data = (islice(r, 0, None) for r in data)
    df = pd.DataFrame(data, columns=cols)            
    # remove unnamedcolumn
    df = removeNullColumn(df)

    return df

def getOpenedFileLists(path, configDF):
    filelist = []
    # os.listdir()方法获取文件夹名字，返回数组
    for root, dirs, files in os.walk(path):
        for file in files:
            filelist.append(os.path.join(root, file))
    # TBD: 支持其他表格格式？
    filelist = list(filter(lambda name: ".xlsx" in name, filelist))

    for file in filelist:

        if fileIsOpen(file):
            # replace替换"["、"]"、" "、"'"
            #file_name = file_name.replace("[", "").replace("]", "").replace("'", "").replace(",", "\n").replace(" ", "")
            openedFileList.append(file)

            if file not in dfDict.keys():
                #df = pd.read_excel(file)
                logger.info("".join([file, '-', "选择载入原始文件"]))
                isToHandle = win32api.MessageBox(0,
                                                 "监测到打开文件{file}。此文件是否开启更新？如果选是，请在初始化完成后再保存文件改动。".format(file=file),
                                                 "初始化选择",
                                                 win32con.MB_YESNO + win32con.MB_SYSTEMMODAL)

                if isToHandle == win32con.IDYES:

                    time_start = time.time()
                    try:
                        wb = openpyxl.load_workbook(file, data_only=True, read_only=True)
                        sheetlist = list(configDF["总表sheet名"])
                        for sheet in sheetlist:
                            df = worksheetToDataFrame(wb, sheet)
                            dfDict[(file,sheet)] = df
                        time_end = time.time()
                        time_spent ='%.2f' %(time_end-time_start)
                        # Test result: 1258 * 55 * 2, 3s
                        logger.info('载入文件{file}成功,耗时:{time_spent}s'.format(file=file,time_spent=time_spent))
                        show_msg("初始化完成",
                                 '载入文件{file}成功,耗时:{time_spent}s'.format(file=file,time_spent=time_spent),
                                 seconds = 5)
                    finally:
                        wb.close()
                else:
                    logger.info("-".join([file, '未载入']))

# 判定文件是否打开
def fileIsOpen(filePath):
    # filepath = C:/Users/Administrator/Desktop/新用户创建.xlsx
    filef = os.path.split(filePath)  # 文件路径和文件名拆开
    excelname = filef[-1]
    excelpath = filef[0]
    hidefilename = excelpath + r"/~$" + excelname  # 拼接出隐藏文件的文件路径
    #print('fileIsOpen:', hidefilename, os.path.exists(hidefilename))
    if os.path.exists(hidefilename):
        return True
    else:
        return False

def check_file_type(f_path):
    '''
    txt = re.compile(r'\.xlsx$')
    if txt.search(f_path) is None:
        # logger.info('只允许xlsx文件')
        print('文件格式格式不正确，只允许xlsx文件')
    if not os.path.exists(f_path):
        print(f_path + ' 文件不存在，请在当前目录检查必要文件')
    if fileIsOpen(f_path):
        print(f_path + ' 文件被占用，请关闭')
    '''
    pass

def addOpenedFile(filePath, configDF):

    filef = os.path.split(filePath)  # 文件路径和文件名拆开
    fname = filef[-1]
    fpath = filef[0]
    if "~$" in fname[0:2]:
        filePath = fpath + "\\" + fname[2:]
        try:
            logger.info(''.join(['打开文件:', filePath]))
            logger.info("".join([filePath, '-', "选择载入原始文件"]))
            isToHandle = win32api.MessageBox(0,
                                             "监测到打开文件{file}。此文件是否开启更新？如果选是，请在初始化完成后再保存文件改动。".format(file=filePath),
                                             "初始化选择", win32con.MB_YESNO + win32con.MB_SYSTEMMODAL )
            if isToHandle == win32con.IDYES:
                try:
                    time_start = time.time()
                    wb = openpyxl.load_workbook(filePath, data_only=True, read_only=True)
                    sheetlist = list(configDF["总表sheet名"])
                    for sheet in sheetlist:
                        df = worksheetToDataFrame(wb, sheet)
                        dfDict[(filePath,sheet)] = df
                    if filePath not in openedFileList:
                        openedFileList.append(filePath)
                    time_end = time.time()
                    time_spent ='%.2f' %(time_end-time_start)
                    # Test result: 1258 * 55, 3s
                    logger.info('载入文件{file}成功,耗时:{time_spent}s'.format(file=filePath,time_spent=time_spent))
                    show_msg("初始化完成",
                             '载入文件{file}成功,耗时:{time_spent}s'.format(file=filePath,time_spent=time_spent),
                             seconds = 5)
                finally:
                    wb.close()
            else:
                logger.info("-".join([filePath, '未载入']))
        except Exception as e:
            logger.error('-'.join([filePath, repr(e)]))
            win32api.MessageBox(0, '-'.join([filePath, repr(e)]), "ERROR", win32con.MB_ICONERROR)
            raise

def removeOpenedFile(filePath):
    
    filef = os.path.split(filePath)  # 文件路径和文件名拆开
    fname = filef[-1]
    fpath = filef[0]

    if "~$" in fname[0:2]:
        # close file
        filePath = fpath + "\\" + fname[2:]
        for i in range(len(openedFileList)):
            if openedFileList[i] == filePath:
                openedFileList.pop(i)
                break
        removedlist = []    
        for key in dfDict.keys():
            dfDictKey = (key[0], key[1])
            removedlist.append(dfDictKey)
        for item in removedlist:
            dfDict.pop(item)
    
        logger.info(''.join(['关闭文件:', filePath]))
        #print('dfDict:', dfDict.keys())

def reloadOpenedFile(modifiedList):

    try:
        time_start = time.time()
        # wb cannot be None
        for (filePath, sheet) in modifiedList.keys():
            dfDict[(filePath,sheet)] = modifiedList[(filePath, sheet)]
            logger.info('重新载入文件-{filePath}-{sheet}'.format(filePath=filePath, sheet=sheet))
        time_end = time.time()
        time_spent = '%.7f'%(time_end-time_start)
        # Test result: 1258 * 55, 0.28s
        logger.info('重新载入文件-{filePath}-{spent}s'.format(filePath=filePath, spent=time_spent))
        
    except Exception as e:
        logger.error("".join([filePath, '-' , repr(e)]))
        win32api.MessageBox(0, filePath, "文件加载错误", win32con.MB_ICONERROR)
        raise

# TBD：TEST多个拆分情况
def getConfigInfo(srcPath):
    # get opening file's config
    configInfo = configDF.loc[configDF['Excel路径'] == srcPath]
    sheetList = list(configInfo['总表sheet名'].drop_duplicates())
    configIndexes = configInfo.index.tolist()
    configDict = {} # key:(sheetName, company, index), value:(savePath, isSummarySet)
    
    for i in range(0, len(configIndexes)):
        configIndex = configIndexes[i]
        
        companyColumn = configInfo.loc[configIndex,'公司列']
        indexColumn = configInfo.loc[configIndex,'索引列']
        splitPath = configInfo.loc[configIndex,'保存路径']
        isSummarySet = configInfo.loc[configIndex,'生成总表']
        isIndexRemain = configInfo.loc[configIndex,'保持索引列']
        sheet = configInfo.loc[configIndex,'总表sheet名']
        
        if (companyColumn, indexColumn) in configDict.keys():
            configDict[(sheet, companyColumn, indexColumn)] += [(splitPath, isSummarySet, isIndexRemain)]
        else:
            configDict[(sheet, companyColumn, indexColumn)] = [(splitPath, isSummarySet, isIndexRemain)]

    return (configDict, sheetList)

def validateNowDF(nowDF):
            
    isValid = True
    errmsgs = ""
    
    duplicatedDF = nowDF[nowDF.duplicated(subset=[idStr])]
    duplicatedDF.dropna(subset=[idStr], inplace=True)
    
    if duplicatedDF.shape[0] > 0:
        isValid = False
        errmsgs = "重复的编号-{id}".format(id=duplicatedDF[idStr].tolist())
    
    return (isValid, errmsgs)

def validateConfig(configDF):
            
    isValid = True
    errmsgs = ""
    
    duplicatedDF = configDF[configDF.duplicated(subset=['Excel路径', '保存路径', '公司列'])]
    if duplicatedDF.shape[0] > 0:
        isValid = False
        duplicatelist = duplicatedDF.index.tolist()
        duplicatelist = [ i+2 for i in duplicatelist]
        errmsgs = "重复配置数据行 - {index}".format(index=duplicatelist)
    configDF.dropna(inplace = True)
    
    return (isValid, errmsgs)

def getDiffDFv2(sheetname, prevDF, nowDF, configInfo):
    
    # spend nearly 0.01s
    
    time_start=time.time()
    
    # create/delete, if company or index is NAN, drop

    createDFs = {} 
    updateDFs = {}
    deleteDFs = {} 
    
    #prevDF = oldDF.fillna('N/A', inplace=False)
    #nowDF = currDF.fillna('N/A', inplace=False)

    for (sheet, companyColumn, indexColumn) in configInfo.keys():

        #print('company,index', companyColumn,indexColumn)
        # nowdf - prevdf
        # Notes: same id, will not create data
        if sheet != sheetname:
            continue

        createDF = pd.concat([nowDF, prevDF, prevDF]).drop_duplicates(
                    subset=[idStr, companyColumn, indexColumn], keep=False)
        
        createAndUpdateDF = pd.concat([nowDF, prevDF, prevDF]).drop_duplicates(
                    keep=False)
        
        updateDF = pd.concat([createAndUpdateDF, createDF, createDF]).drop_duplicates(
                    keep=False)
        
        deleteDF = pd.concat([prevDF, nowDF, nowDF]).drop_duplicates(
                    subset=[idStr, companyColumn, indexColumn], keep=False)
        
        createDF.dropna(subset=[idStr, companyColumn, indexColumn], inplace=True)
        updateDF.dropna(subset=[idStr, companyColumn, indexColumn], inplace=True)
        deleteDF.dropna(subset=[idStr, companyColumn, indexColumn], inplace=True)
        
        if len(createDF) >= 1:
            createDFs[(companyColumn, indexColumn)] = createDF
        if len(updateDF) >= 1:
            updateDFs[(companyColumn, indexColumn)] = updateDF
        if len(deleteDF) >= 1:
            deleteDFs[(companyColumn, indexColumn)] = deleteDF
        
    #print('createDFs', createDFs)
    #print('updateDFs', updateDFs)
    #print('deleteDFs', deleteDFs)
    
    time_end=time.time()
    time_spent = '%.2f' %(time_end-time_start)

    # spend 0.15s
    logger.info("比较完毕-{time_spent}s".format(time_spent=time_spent))
    return (createDFs, updateDFs, deleteDFs)

def updateCell(saveWB, group, sheetname):

    saveWS = saveWB[sheetname]
    sheetDF = worksheetToDataFrame(saveWB, sheetname, ws=saveWS)

    seqColumn = -1
    '''
    for col in saveWS.iter_rows(min_row=1, max_row=1, values_only=True):
        header = list(col)
        while None in header:
            header.remove(None)
    '''
    header = list(sheetDF.columns)
    group = group.reindex(columns=list(sheetDF.columns), fill_value='\\')
    if "序号" in header:
        seqColumn = (header.index("序号"))
    for newIndex, newRow in group.iterrows():
        groupid = newRow[idStr]
        oldRows = sheetDF[sheetDF[idStr] == groupid] # df with one row

        # although loop, actually just one Series, otherwise error
        if oldRows.shape[0] == 1:
            for oldIndex, oldRow in oldRows.iterrows():
                column = 0
                for key_value in oldRow.items():
                    # no need to compare sequence:
                    if seqColumn != -1 and column == seqColumn:
                        column += 1
                        continue
                    columnName = key_value[0]
                    oldValue = key_value[1]
                    # columnname, unamed or nan, Test:ok, group has removed null cols
                    newValue = newRow[columnName]
                    # ids are the same
                    if pd.isnull(oldValue) and pd.isnull(newValue):
                        column += 1
                        continue
                    # Notes: format should be same
                    elif oldValue != newValue:
                        #print('cell', oldIndex, column)
                        # Notes: cell(): if value is not None, then can be updated
                        if newValue == None:
                            newValue = ''
                        # Notes: index + 2 considering header
                        cell = saveWS.cell(oldIndex+2, column+1, newValue)
                        cell.font = cellFont
                        cell.border = cellBorder
                        cell.alignment= cellAlignment
                        
                    column += 1
        else:
            logger.error("拆分表{sheetname}-{groupid}重复".format(sheetname=sheetname, groupid=groupid))
            win32api.MessageBox(0,
                                "拆分表{sheetname}-{groupid}重复，请检查该文件！".format(sheetname=sheetname, groupid=groupid),
                                "ERROR",
                                win32con.MB_ICONERROR)


def synchorizeDF(summarySheetname, createDFs, updateDFs, deleteDFs, configInfo):
    
    isSync = False

    if (createDFs != None) and (len(createDFs.keys()) > 0):
        isSync = True
    elif (updateDFs != None) and (len(updateDFs.keys()) > 0):
        isSync = True
    elif (deleteDFs != None) and (len(deleteDFs.keys()) > 0):
        isSync = True
    
    # configInfo:(company, index), info for the srcPath
    
    openWB = {}
    openWriter = {}
    
    # companyColumn: column name
    try:
        if isSync:
            time_start=time.time()
            for (sheet, companyColumn, indexColumn) in configInfo.keys():
                if sheet != summarySheetname:
                    continue
                isCreate, isUpdate, isDelete = False, False, False
                configs = configInfo[(sheet, companyColumn, indexColumn)]
                if (companyColumn, indexColumn) in deleteDFs.keys():
                    deleteDF = deleteDFs[(companyColumn, indexColumn)]
                    isDelete = True
                    #print('deleteData', deleteDF)
                if (companyColumn, indexColumn) in updateDFs.keys():
                    updateDF = updateDFs[(companyColumn, indexColumn)]
                    isUpdate = True
                    #print('UpdateData', updateDF)
                if (companyColumn, indexColumn) in createDFs.keys():
                    createDF = createDFs[(companyColumn, indexColumn)]
                    isCreate = True
                    #print('createData', createDF)            
                
                for config in configs:
                    splitPath = config[0]
                    isSummarySet = config[1]
                    isIndexRemain = config[2]
                    
                    # (company, index): company, index value
        
                    # Notes:delete should be prior to create, considering update index            
        
                    # delete
                    # Notes: 这家公司数据删光，不会删，考虑到误删
                    # use openpyxl to delete
                    if isDelete:
                        for (company, index), group in deleteDF.groupby([companyColumn, indexColumn]):
                            savePath = splitPath + "\\" + company + ".xlsx"
                            if not os.path.isfile(savePath):
                                show_msg("找不到文件",
                                         "删除记录失败，{file}不存在。".format(file=savePath),
                                         seconds = leftSeconds)
                                continue
                                
                            if savePath in openWB.keys():
                                saveWB = openWB[savePath]
                                saveWriter = openWriter[savePath]
                            else:
                                saveWB = openpyxl.load_workbook(savePath, data_only=False)
                                saveWriter = pd.ExcelWriter(savePath, engine="openpyxl")
                                openWB[savePath] = saveWB
                                openWriter[savePath] = saveWriter
                            saveWriter.book = saveWB
                            saveWriter.sheets = dict((ws.title, ws) for ws in saveWB.worksheets)
                            # in diff, already drop index = none
                            if index == '\\':
                                index = '空'
                            index = str(index)
                        
                            # column deleted if not isIndexRemain
                            if not isIndexRemain:
                                del group[companyColumn]
                                del group[indexColumn]
                            logger.info('删除数据-{group}'.format(group=group))
                            
                            groupIDs = list(group[idStr])
                            
                            if isSummarySet:
                                sheetname1 = saveWB.sheetnames[0]
                                # use pandas delete, stye cannot remain?
                                saveWS = saveWB[sheetname1]
                                for col in saveWS.iter_rows(min_row=1, max_row=1, values_only=True):
                                    header = list(col)
                                idColumn = (header.index("编号"))
                                for groupID in groupIDs:
                                    
                                    subdata = np.array(list(saveWS.values))
                                    subdata = subdata[1:,idColumn:idColumn+1]
                                    temp = np.argwhere(subdata == groupID)
                                    if len(temp > 0):
                                        deleteRow = temp[0][0]
                                        #print('deleteRowNumber:', deleteRow)
                                        # 2: header, 1-base
                                        saveWS.delete_rows(deleteRow + 2)
                                    else:
                                        logger.warning("-".join([savePath, sheetname1, groupID, "找不到"]))
                                        win32api.MessageBox(0, "-".join([savePath, sheetname1, groupID]), 
                                                            "拆分总表记录丢失",
                                                            win32con.MB_ICONWARNING)
                                    
                            saveWS = saveWB[index]
                            for col in saveWS.iter_rows(min_row=1, max_row=1, values_only=True):
                                header = list(col)
                            idColumn = (header.index("编号"))
                            for groupID in groupIDs:
                                subdata = np.array(list(saveWS.values))
                                subdata = subdata[1:,idColumn:idColumn+1]
                                temp = np.argwhere(subdata == groupID)
                                if len(temp > 0):
                                    deleteRow = temp[0][0]
                                    #print('deleteRowNumber:', deleteRow)
                                    # 2: header, 1-base
                                    saveWS.delete_rows(deleteRow + 2)
                                else:
                                    logger.warning("-".join([savePath, index, groupID, "找不到"]))
                                    win32api.MessageBox(0, "-".join([savePath, index, groupID]), 
                                                        "拆分分表记录丢失",
                                                        win32con.MB_ICONWARNING)
                                    
                            
                            # if no rows in subsheet, delete subsheet
                            subdata = np.array(list(saveWS.values))
                            if len(subdata) == 1:
                                del saveWB[index]
    
                    # create
                    # 新加公司（文件不存在,使用拆分脚本(单个公司操作),此脚本不涉及
                    # 防止公式丢失,读,data_only = True,写,data_only=False
                    if isCreate:
                        for (company, index), group in createDF.groupby([companyColumn, indexColumn]):
                            savePath = splitPath + "\\" + company + ".xlsx"
                            if not os.path.isfile(savePath):
                                # 公司不能新加excel，考虑到频繁添加excel
                                logger.warning("{savePath}找不到，不支持添加新公司！".format(savePath=savePath))
                                win32api.MessageBox(0, "{savePath}找不到，不支持添加新公司！".format(savePath=savePath), 
                                                    "无效公司名",
                                                    win32con.MB_ICONWARNING)
                                continue
                            savePath = splitPath + "\\" + company + ".xlsx"
                            if savePath in openWB.keys():
                                saveWB = openWB[savePath]
                                saveWriter = openWriter[savePath]
                            else:
                                # Notes: 读公式,data_only = True,写,data_only = False
                                saveWB = openpyxl.load_workbook(savePath, data_only=False)
                                saveWriter = pd.ExcelWriter(savePath, engine="openpyxl")
                                openWB[savePath] = saveWB
                                openWriter[savePath] = saveWriter
                            saveWriter.book = saveWB
                            saveWriter.sheets = dict((ws.title, ws) for ws in saveWB.worksheets)
                            
                            if index == '\\':
                                index = '空'
                            index = str(index)
                            
                            # column deleted if not isIndexRemain
                            if not isIndexRemain:
                                del group[companyColumn]
                                del group[indexColumn]
                            logger.info('添加数据-{group}'.format(group=group))
                            # append data
                            # TBD: test subsheet disorder?
                            if isSummarySet:
                                summarySheet = saveWB.sheetnames[0]
                                saveWS = saveWB[summarySheet]
        
                                for col in saveWS.iter_rows(min_row=1, max_row=1, values_only=True):
                                    header = list(col)
                                    while None in header:
                                        header.remove(None)
                                group = group.reindex(columns=header, fill_value='\\')
                                
                                if "序号" in header:
                                    group.loc[:, "序号"] = '=ROW()-1'
                                minRow = saveWS.max_row + 1
                                maxRow = minRow + group.shape[0] - 1
                                for r in dataframe_to_rows(group, index=False, header=False):
                                    saveWS.append(r)
                                for r in saveWS.iter_rows(min_row=minRow, max_row=maxRow, values_only=False):
                                    for cell in r:
                                        #设置表头
                                        cell.font = cellFont
                                        cell.border = cellBorder
                                        cell.alignment = cellAlignment
        
                            if index in saveWB.sheetnames:
                                # if subsheet exists, add to subsheet
                                saveWS = saveWB[index]
                                for col in saveWS.iter_rows(min_row=1, max_row=1, values_only=True):
                                    header = list(col)
                                    while None in header:
                                        header.remove(None)
                                group = group.reindex(columns=header, fill_value='\\')
                                
                                if "序号" in header:
                                    group.loc[:, "序号"] = '=ROW()-1'
                                minRow = saveWS.max_row + 1
                                maxRow = minRow + group.shape[0] - 1
                                for r in dataframe_to_rows(group, index=False, header=False):
                                    saveWS.append(r)
                                for r in saveWS.iter_rows(min_row=minRow, max_row=maxRow, values_only=False):
                                    for cell in r:
                                        #设置表头
                                        cell.font = cellFont
                                        cell.border = cellBorder
                                        cell.alignment = cellAlignment
                            else:
                                # use first sheet header
                                firstSheet = saveWB.sheetnames[0]
                                saveWS = saveWB[firstSheet]
                                for col in saveWS.iter_rows(min_row=1, max_row=1, values_only=True):
                                    header = list(col)
                                    while None in header:
                                        header.remove(None)
                                group = group.reindex(columns=header, fill_value='\\')
                                if "序号" in header:
                                    group.loc[:, "序号"] = '=ROW()-1'
                                # if subsheet not exists, create new sheet
                                saveWS = saveWB.create_sheet(index)
                                for r in dataframe_to_rows(group, index=False, header=True):
                                    saveWS.append(r)
                                count = 1
                                for r in saveWS:
                                    for cell in r:
                                        #设置表头
                                        cell.font = cellFont
                                        cell.border = cellBorder
                                        if count == 1:
                                            cell.alignment = Alignment(
                                                horizontal='center',  # 水平对齐，可选general、left、center、right、fill、justify、centerContinuous、distributed
                                                vertical='center',  # 垂直对齐， 可选top、center、bottom、justify、distributed
                                                wrapText=True) #自动换行
                                        else:
                                            cell.alignment = cellAlignment
                                            
                                    count += 1
                        
                    # update
                    # only update data without id, company, index
                    # use openpyxl and pandas
                    if isUpdate:
                        for (company, index), group in updateDF.groupby([companyColumn, indexColumn]):
                            savePath = splitPath + "\\" + company + ".xlsx"
                            if savePath in openWB.keys():
                                saveWB = openWB[savePath]
                                saveWriter = openWriter[savePath]
                            else:
                                saveWB = openpyxl.load_workbook(savePath, data_only=False)
                                saveWriter = pd.ExcelWriter(savePath, engine="openpyxl")
                                openWB[savePath] = saveWB
                                openWriter[savePath] = saveWriter
                            saveWriter.book = saveWB
                            saveWriter.sheets = dict((ws.title, ws) for ws in saveWB.worksheets)
                            
                            if index == '\\':
                                index = '空'
                            index = str(index)
                        
                            # column deleted if not isIndexRemain
                            if not isIndexRemain:
                                del group[companyColumn]
                                del group[indexColumn]
                            logger.info('更新数据-{company}-{index}-{group}'.format(company=company,index=index,group=group))
                            # use pandas to find cells to be updated
                            if isSummarySet:
                                sheetname = saveWB.sheetnames[0]
                                updateCell(saveWB, group, sheetname)
                            # update subsheet
                            updateCell(saveWB, group, index)
    except PermissionError as e:
        isSync = False
        logger.error(repr(e))
        win32api.MessageBox(0, repr(e), "访问权限错误", win32con.MB_ICONERROR)
    except Exception as e:
        isSync = False
        logger.error(repr(e))
        win32api.MessageBox(0, repr(e), "ERROR", win32con.MB_ICONERROR)
        raise
    finally:
        # save, TBD:ROLL BACK
        for savePath in openWriter.keys():
            saveWriter = openWriter[savePath]
            # Notes: if completed, writer needs to be closed, otherwise file damages
            # Use workbook.save() instead of using an `ExcelWriter`.
            #saveWB.save(excelpath + excelname)
            saveWB.close()
            saveWriter.close()
        if isSync:
            time_end=time.time()
            time_spent = '%.2f' %(time_end-time_start)
            logger.info("同步数据花费-{time_spent}s".format(time_spent=time_spent))

        return isSync

class ScriptEventHandler(FileSystemEventHandler):
    
    
    def __init__(self):
        FileSystemEventHandler.__init__(self)

    '''
    # 文件移动
    def on_moved(self, event):
        if event.is_directory:
            print("directory moved from {src_path} to {dest_path}".format(src_path=event.src_path,
                                                                          dest_path=event.dest_path))
        else:
            print(
                "file moved from {src_path} to {dest_path}".format(src_path=event.src_path, dest_path=event.dest_path))
    '''
    
    # 文件新建
    def on_created(self, event):
        if event.is_directory:
            pass
            #("directory created:{file_path}".format(file_path=event.src_path))
        else:
            #print("file created:{file_path}".format(file_path=event.src_path))
            addOpenedFile(event.src_path, configDF)
            
    # 文件删除
    def on_deleted(self, event):
        if event.is_directory:
            pass
            #print("directory deleted:{file_path}".format(file_path=event.src_path))
        else:
            #print("file deleted:{file_path}".format(file_path=event.src_path))
            removeOpenedFile(event.src_path)


    # 文件修改
    def on_modified(self, event):
        
        filePath = event.src_path
        # may modify filename like 357DCA00
        if ".xlsx" not in os.path.basename(filePath):
            return
        
        global saveCount
        # save file, triggle 2 times for on_modified (windows)
        if saveCount >= 1:
            saveCount = 0
            return
        else:
            saveCount = saveCount + 1

        #print('saveCount:', saveCount)
        if event.is_directory:
            pass
            #print("directory modified:{file_path}".format(file_path=event.src_path))
        else:
            isloadedOldDF = False
            for key in dfDict.keys():
                if filePath == key[0]:
                    isloadedOldDF = True
                    break
            if isloadedOldDF:
                isChanged = False
                wb = None
                # 有变化的文件, 需要重载
                modifiedList = {} # key: (filePath, sheetname), value:dataframe
                try:
                    time_start=time.time()
                    createDFs, updateDFs, deleteDFs = None, None, None
                    
                    logger.info("修改的文件:{file_path}".format(file_path=filePath))
                    
                    # read index column, isSummarySet from configInfo
                    configInfo, sheetList = getConfigInfo(event.src_path)

                    if configInfo == {} or len(sheetList) == 0:
                        logger.error("".join([filePath, '-', "该文件配置信息不存在"]))
                        win32api.MessageBox(0, "该文件配置信息不存在！请检查配置文件。", "配置错误", win32con.MB_ICONERROR)
                        # TBD raise filenotfound? 
                    #print (configInfo)

                    # TBD:一张excel, 多个sheet
                    time_start_read=time.time()
                    wb = openpyxl.load_workbook(filePath, data_only=True, read_only=True)
                    for sheet in sheetList:
                        # load df
                        # read original file
                        oldDF = dfDict[(filePath,sheet)]
                        # read current file
                        nowDF = worksheetToDataFrame(wb, sheet)
                        time_end_read=time.time()
                        time_spent_read = '%.2f' %(time_end_read-time_start_read)
                        logger.info("-".join([filePath,'读取完毕', time_spent_read,'s']))
                        # Validation and Warning msgs
                        isValid, errmsgs = validateNowDF(nowDF)
                        
                        # diff file
                        if isValid:
                            createDFs, updateDFs, deleteDFs = getDiffDFv2(sheet, oldDF, nowDF, configInfo)
                        else:
                            logger.error(errmsgs)
                            win32api.MessageBox(0, errmsgs, "数据错误", win32con.MB_ICONERROR)
                            return
        
                        # 3.synchronize
                        # company, index: columnName
                        #return
                        isChanged = synchorizeDF(sheet,
                                                 createDFs,
                                                 updateDFs,
                                                 deleteDFs,
                                                 configInfo)
                        if isChanged:
                            # to reload
                            modifiedList[(filePath, sheet)] = nowDF

                except Exception as e:
                    logger.error("".join([filePath, '-', repr(e)]))
                    win32api.MessageBox(0, repr(e), "ERROR", win32con.MB_ICONERROR)
                    raise
                finally:
                    if wb != None:
                        wb.close()
                    time_end=time.time()
                    time_spent = '%.2f' %(time_end-time_start)
                    # 若有修改，reload该文件
                    if isChanged:
                        show_msg("同步成功",
                                 "同步数据共耗时{time_spent}s。".format(time_spent=time_spent),
                                 seconds = leftSeconds)
                        reloadOpenedFile(modifiedList)
                        time_end=time.time()
                        time_spent = '%.2f' %(time_end-time_start)
                        logger.info("数据更新-{file_path}-{time_spent}s".format(file_path=filePath, time_spent=time_spent))
                    else:
                        logger.info("无数据更新-{file_path}-{time_spent}s".format(file_path=filePath, time_spent=time_spent))
                        show_msg("无数据更新",
                                 filePath,
                                 seconds = leftSeconds)
                    
                

 
if __name__ == "__main__":

    # 创建一个logger
    logger = setLogger()

    # 获取索引列信息配置表
    try:
        time_start = time.time()
        configDF = pd.read_excel(configPath)
        (isValid, errmsgs) = validateConfig(configDF)
        if not isValid:
            logger.error(errmsgs)
            win32api.MessageBox(0, errmsgs, "配置文件错误", win32con.MB_ICONERROR)
            raise Exception
        time_end = time.time()
        time_spent = "%.2f" %(time_end-time_start)
        logger.info("{configPath}-加载成功-耗时{time_spent}s".format(configPath=configPath,time_spent=time_spent))
        show_msg("配置文件读取","{configPath}加载成功，耗时{time_spent}s".format(configPath=configPath,time_spent=time_spent),
                 seconds=5)
        
    except Exception as e:
        logger.error("".join([configPath, '-', repr(e)]))
        win32api.MessageBox(0, "配置文件读取错误！请检查后重启程序。", "ERROR",
                            win32con.MB_ICONERROR)
        raise

    # 获取该目录下所有打开文件，启动时获取一次
    fileList = getOpenedFileLists(userPath, configDF)

    event_handler1 = ScriptEventHandler()
    observer = Observer()
    watch = observer.schedule(event_handler1,
                              path=userPath,
                              recursive=True)

    observer.start()
    try:
        while True:
            time.sleep(1)   
    except KeyboardInterrupt:
        observer.stop()
    observer.join()