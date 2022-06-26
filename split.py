# -*- coding: utf-8 -*-
"""
Created on Sat Jun 11 19:14:49 2022

Split the excel into new excels according to time and coporation

@author: Jane Huang
"""

import tkinter as tk
import time as tm
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles import Border
from openpyxl.styles import Side
from os.path import isfile
from os.path import isdir
from os.path import abspath
from tkinter.filedialog import askopenfilename
from tkinter.filedialog import askdirectory
from tkinter import scrolledtext

isDebug = False #使用固定填入
showProgress = False
# 想要生成的excel文件个数，如果想要全部文件生成，设置isSetFileNums为False
isSetFileNums = False
setFileNums = 2
# 想要单独生成的公司文件，如果想要全部公司生成，设置isSetFileNums为False
isSetCompanyName = False
setCompanyName = "上化氟"
# 拆分信息维护文件
configPath = 'C:\\Users\\Lenovo\Desktop\\make money\\拆分\\config\\configuration.xlsx'

#设置excel样式
cellFont = Font(
    name="宋体",   # 字体
    size=9,         # 字体大小
    color="000000",  # 字体颜色，用16进制rgb表示
    bold=False,       # 是否加粗，True/False
    italic=False,     # 是否斜体，True/False
    #strike=False,     # 是否使用删除线，True/False
    #underline=None,  # 下划线, 可选'singleAccounting', 'double', 'single', 'doubleAccounting'
)
cellAlignment = Alignment(
    horizontal='center',  # 水平对齐，可选general、left、center、right、fill、justify、centerContinuous、distributed
    vertical='center',  # 垂直对齐， 可选top、center、bottom、justify、distributed
    #text_rotation=0,  # 字体旋转，0~180整数
    #wrap_text=False,  # 是否自动换行
    #shrink_to_fit=False,  # 是否缩小字体填充
    #indent=0,  # 缩进值
)
cellBorder = Border(bottom=Side(style='thin'),
                    right=Side(style='thin'))

def setExcelStyle(sheet, start_row):
    #设置excel样式
    #print("===========设置表样式中=========\n")
    # set header
    if start_row == 1:
        for row in sheet.iter_rows(min_row=1, max_row = 1,values_only=False):
            for cell in row:
                #设置表头
                cell.font = cellFont
                cell.border = cellBorder
                cell.alignment = Alignment(
                    horizontal='center',  # 水平对齐，可选general、left、center、right、fill、justify、centerContinuous、distributed
                    vertical='center',  # 垂直对齐， 可选top、center、bottom、justify、distributed
                    wrapText=True) #自动换行
        start_row += 1

    for row in sheet.iter_rows(min_row=start_row, values_only=False):
        for cell in row:
            #设置表头
            cell.font = cellFont
            cell.border = cellBorder
            cell.alignment = cellAlignment

def selectFile():
    path_ = askopenfilename() #使用askdirectory()方法返回文件夹的路径
    if path_ == "":
        path1.get() #当打开文件路径选择框后点击"取消" 输入框会清空路径，所以使用get()方法再获取一次路径
    else:
        path_ = path_.replace("/", "\\")  # 实际在代码中执行的路径为“\“ 所以替换一下
        path1.set(path_)
    if txt1.get() != "" and path_ != "":
        #print("delete")
        txt1.delete(0, 'end')
    txt1.insert('insert', path_)

def selectPath():
    path_ = askdirectory() #使用askdirectory()方法返回文件夹的路径
    if path_ == "":
        path2.get() #当打开文件路径选择框后点击"取消" 输入框会清空路径，所以使用get()方法再获取一次路径
    else:
        path_ = path_.replace("/", "\\")  # 实际在代码中执行的路径为“\“ 所以替换一下
        path2.set(path_)
    if txt2.get() != "" and path_ != "":
        #print("delete")
        txt2.delete(0, 'end')
    txt2.insert('insert', path_)


def changeColumnToNum(column):
    index = 0
    column = column.strip().upper()
    if len(column) > 1:
        base = 1
        for char in column[::-1]:
            index += base * (ord(char) - ord('A') + 1)
            base = base * 23
            #print(char,index, base)
        index -= 1
    else:
        index = ord(column) - ord('A')
    return index
    
def outputConfig(config):
    
    isCreate = False
    if isfile(configPath):
        wb = openpyxl.load_workbook(configPath)
    else:
        wb = openpyxl.Workbook()
        isCreate = True
    ws = wb.active
    header = ['Excel路径', '保存路径', '公司列', '索引列', '生成总表', '保持索引列', '总表sheet名']
    if isCreate:
        ws.append(header)
    ws.append(config)
    wb.save(configPath)

def splitExcel():
    time_start = tm.time()
    #读取excel文件
    '''
        下列数据仅用于调试
    '''
    
    if isDebug:
        txt1.insert(0, "C:\\Users\\Lenovo\Desktop\\make money\\拆分\\test\\汇总2.xlsx")
        txt1.update()
        txtSheet.insert(0, "总表")
        txtSheet.update()
        txt2.insert(0, "C:\\Users\\Lenovo\\Desktop\\make money\\拆分\\result\\花名册岗位")
        txt2.update()
        txtIndex.insert(0, 'K')
        txtIndex.update()
        txtCompany.insert(0, 'C')
        txtCompany.update()
        txtSpecial.insert(0, 'F,N,O,W')
        txtSpecial.update()
    
    msgs = ""
    excelPath = txt1.get()
    sheetName = txtSheet.get()
    dirPath = txt2.get()
    timeColumn = changeColumnToNum(txtIndex.get())
    companyColumn = changeColumnToNum(txtCompany.get())
    specialColumns = txtSpecial.get()
    
    if excelPath != "" and dirPath != "" and sheetName != "":
        # excel文件存在///
        if not isfile(excelPath):
            msgs = ''.join([msgs, "文件", excelPath, "不存在\n"])
        # 目的路径存在
        if not isdir(dirPath):
            msgs = ''.join([msgs, "目录", dirPath, "不存在\n"])
        if isfile(excelPath) and isdir(dirPath):
            outNums = 0
            try:
                #start = time.time()
                myBook = openpyxl.load_workbook(excelPath, read_only=True)
                mySheet = myBook[sheetName]
                #按行获取excel(mySheet)的单元格数据(myRange)
                myRange = list(mySheet.values)
                #创建空白字典(myDict)
                myDict={}
                #删除表头company和index列
                myRange[0] = list(myRange[0])
                #记录原始列名，后面可能会被删除
                companyColumnName = myRange[0][companyColumn]
                timeColumnName = myRange[0][timeColumn]
                if isIndexRemain.get() != 1:
                    myRange[0].pop(timeColumn)
                    if companyColumn > timeColumn:
                        myRange[0].pop(companyColumn - 1)
                    else:
                        myRange[0].pop(companyColumn)
                #添加序号列表头
                if "序号" in myRange[0]:
                    # 设序号,本身文件要有,没有会抛错
                    seqColumn = myRange[0].index("序号")
                    myRange[0][seqColumn] =  "序号"
                
                #从excel(myRange)的第2行开始循环(到最后一行)
                #seq = 1
                #处理加载数据
                print("=================开始解析文件\n")
                indexes = []
                if specialColumns.strip() != '':
                    specialColumns = specialColumns.split(',')
                    for specialColumn in specialColumns:
                        indexes.append(changeColumnToNum(specialColumn))
                
                for myRow in myRange[1:]:
                    #拆分依据的列名(可修改),按公司和时间拆
                    time = str(myRow[timeColumn])
                    company = myRow[companyColumn]
                    if isSetCompanyName and company != setCompanyName:
                        continue
                    #待输入工号，先跳过
                    if time == '\\' or time.strip() == '':
                        time = "空"
                    myRow = list(myRow)
                    #设置特殊字符列
                    for index in indexes:
                        myRow[index] = str(myRow[index])

                    dictKey = (company, time)

                    # 添加数据
                    if dictKey in myDict.keys():
                        '''
                        if isSeqSet:
                            seq = len(myDict[dictKey]) + 1
                            seq += 1
                        '''
                        myDict[(company,time)] += [myRow]
                    else:
                        '''
                        if isSeqSet:
                            seq = 1
                        '''
                        myDict[(company,time)]=[myRow]
                    #删除每一行company和index列
                    if isIndexRemain.get() != 1:
                        myRow.pop(timeColumn)
                        if companyColumn > timeColumn:
                            myRow.pop(companyColumn - 1)
                        else:
                            myRow.pop(companyColumn)
                    # 设置序号列，注：写死了表头叫序号
                    if "序号" in myRange[0]:
                        seqColumn = myRange[0].index("序号")
                        myRow[seqColumn] = '=ROW()-1'
                    #print("================="+ seq + "\n")
                myBook.close()
                
                #循环字典(myDict)的成员
                prevCompany = ""
                prevNewBook = None
                prevTime = ""
                prevNewSheet = None
                myDictSorted = sorted(myDict.keys(), key = lambda x:(x[0], x[1]))
                start_row_sub = 1
                start_row_default = 1
                
                for key in myDictSorted:
                    if isSetFileNums and outNums >= setFileNums:
                        break
                    value = myDict[key]
                    company = key[0]
                    time = key[1]
                    #创建新工作簿(myNewBook)
                    if prevCompany != company:
                        myNewBook=openpyxl.Workbook()
                        prevTime = ""
                        del myNewBook["Sheet"]
                        if isSummarySet.get() == 1:
                            defaultSheet = myNewBook.create_sheet("总表")
                            defaultSheet.append(myRange[0])
                            start_row_default = 1
                    else:
                        myNewBook = prevNewBook
                    if prevTime != time:
                        #生成新sheet
                        myNewSheet = myNewBook.create_sheet(time)
                        #在新工作表(myNewSheet)中添加表头
                        myNewSheet.append(myRange[0])
                        #print ("生成sheet：", time, "\n")
                        start_row_sub = 1
                    else:
                        myNewSheet = prevNewSheet
                    
                    #print("===========添加数据中=========\n")
                    for data in value:
                        #添加总表数据
                        if isSummarySet.get() == 1:
                            defaultSheet.append(data)
                        #在新工作表(myNewSheet)中添加数据
                        myNewSheet.append(data)

                    # 设置新加数据格式
                    if isSummarySet.get() == 1:
                        setExcelStyle(defaultSheet, start_row_default)
                    setExcelStyle(myNewSheet, start_row_sub)
                    #保存各个Excel文件
                    if prevCompany != '' and prevCompany != company:
                        # 前面公司的已完成，保存
                        outNums += 1
                        savePath = dirPath + '\\' + prevCompany + '.xlsx'
                        prevNewBook.save(savePath)
                        msgs = ''.join(["===", msgs, "生成文件：", savePath, "完成===\n"])
                        print("===生成文件：" + savePath + "完成===\n")
                    else:
                        myPath = dirPath + '\\' + company + '.xlsx'
                        #msgs = ''.join([">>>", msgs, "生成文件：", myPath, "中<<<\n"])
                        if showProgress:
                            print(">>>生成文件：" + myPath + "中<<<\n")
                    if isSummarySet.get() == 1:
                        start_row_default = defaultSheet.max_row + 1
                    start_row_sub = myNewSheet.max_row + 1
                    prevCompany = company
                    prevTime = time
                    prevNewBook = myNewBook
                    prevNewSheet = myNewSheet
                
                # 最后一个excel
                if prevNewBook != None:
                    outNums += 1
                    savePath = dirPath + '\\' + prevCompany + '.xlsx'
                    prevNewBook.save(savePath)

                    msgs = ''.join(["===", msgs, "生成文件：", savePath, "完成===\n"])
                    print("===生成文件：" + savePath + "完成===\n")
                    msgs = ''.join([msgs, "\n生成结束！\n"])
                    print("\n生成结束!\n")
                    config = []
                    excelPath = txt1.get()
                    sheetName = txtSheet.get()
                    dirPath = txt2.get()
                    #specialColumns = txtSpecial.get()
                    config.append(excelPath)
                    config.append(dirPath)
                    config.append(companyColumnName)
                    config.append(timeColumnName)
                    config.append(isSummarySet.get())
                    config.append(isIndexRemain.get())
                    config.append(sheetName)
                    outputConfig(config)
            except Exception as e:
                msgs = ''.join([msgs, "error:", repr(e)])
                raise
            finally:
                time_end = tm.time()
                '''
                    测试结果：
                        个税，31212*20数据，显示部分进度条，67.81s, 68个文件; 
                        花名册，1253*55数据，显示部分进度条，21.60s, 50个文件;
                        pandas groupby, python dict 18ms
                '''
                msgs = ''.join([msgs, '本次运行时间：', str(time_end-time_start) ,'s, 共创建', str(outNums), '个文件\n'])
                print('本次运行时间：',time_end-time_start,'s, 共创建', outNums, '个文件\n')
    else:
        msgs = "excel拆分，目的路径和sheet名不能为空\n"
    scroll.insert(tk.END, msgs)
    scroll.update()


def clearMsgs():
    if scroll.get('1.0', 'end-1c') != "":    
        scroll.delete('1.0', 'end-1c')

if __name__ == "__main__":
    wnd = tk.Tk()
    wnd.title('tools')
    wnd.geometry('420x450')
    wnd.resizable(height=False, width=False)
    
    lbl1 = tk.Label(wnd, text='汇总Excel路径')
    lbl1.grid(row=0, column=0, sticky='E')
    txt1 = tk.Entry(wnd)
    txt1.grid(row=0, column=1, sticky='E')
    path1 = tk.StringVar()
    path1.set(abspath("."))
    tk.Button(wnd,
              text="选择文件",
              command=selectFile).grid(row=0, column=2, sticky='W')

    lblSheet = tk.Label(wnd, text='Sheet')
    lblSheet.grid(row=1, column=0, sticky='E')
    txtSheet = tk.Entry(wnd)
    txtSheet.grid(row=1, column=1, sticky='E')
    lblSheet = tk.Label(wnd, text='(需被拆分的Sheet名字,不可为空)')
    lblSheet.grid(row=1, column=2, sticky='W')
    
    lbl2 = tk.Label(wnd, text='保存路径')
    lbl2.grid(row=2, column=0, sticky='E')
    txt2 = tk.Entry(wnd)
    txt2.grid(row=2, column=1, sticky='E')
    path2 = tk.StringVar()
    path2.set(abspath("."))
    tk.Button(wnd,
              text="选择路径",
              command=selectPath).grid(row=2, column=2, sticky='W')

    lblCompany = tk.Label(wnd, text='公司所在列')
    lblCompany.grid(row=3, column=0, sticky='E')
    txtCompany = tk.Entry(wnd)
    txtCompany.grid(row=3, column=1, sticky='E')
    lblCompany = tk.Label(wnd, text='(例如T,不可为空)')
    lblCompany.grid(row=3, column=2, sticky='W')

    lblIndex = tk.Label(wnd, text='索引列')
    lblIndex.grid(row=4, column=0, sticky='E')
    txtIndex = tk.Entry(wnd)
    txtIndex.grid(row=4, column=1, sticky='E')
    lblIndex = tk.Label(wnd, text='(例如A,不可为空)')
    lblIndex.grid(row=4, column=2, sticky='W')

    # 如果纯数字数据大于？位
    lblSpecial = tk.Label(wnd, text='特殊字符列')
    lblSpecial.grid(row=5, column=0, sticky='E')
    txtSpecial = tk.Entry(wnd)
    txtSpecial.grid(row=5, column=1, sticky='E')
    lblSpecial = tk.Label(wnd, text='>11位(例如M,N,AN,没有则空)')
    lblSpecial.grid(row=5, column=2, sticky='W')
    
    #是否需要总表
    isSummarySet = tk.IntVar()
    lblSummary = tk.Label(wnd, text='生成总表')
    lblSummary.grid(row=6, column=0, sticky='E')
    ckSummary = tk.Checkbutton(wnd, variable=isSummarySet, command='')
    ckSummary.grid(row=6, column=1, sticky='W')

    #是否保留拆分依据
    isIndexRemain = tk.IntVar()
    lblIndexRemain = tk.Label(wnd, text='保留拆分依据列')
    lblIndexRemain.grid(row=7, column=0, sticky='E')
    ckIndexRemain = tk.Checkbutton(wnd, variable=isIndexRemain, command='')
    ckIndexRemain.grid(row=7, column=1, sticky='W')
    
    btn3 = tk.Button(wnd, text='清除消息框', command=clearMsgs)
    btn3.grid(row=8, column=1, sticky='W')
    btn1 = tk.Button(wnd, text='拆分Excel', command=splitExcel)
    btn1.grid(row=8, column=1, sticky='E')
    
    # Message
    scroll = scrolledtext.ScrolledText(wnd,width=40,height=13,font=('黑体',10))
    scroll.grid(row=9, column=0, columnspan=3, pady = 5 , padx = 5 )
    
    wnd.mainloop()
